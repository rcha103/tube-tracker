import io
import json
import mimetypes
import os
import re
import sqlite3
import uuid
from datetime import datetime, timezone

from flask import Flask, abort, jsonify, request, send_file, send_from_directory
from werkzeug.utils import secure_filename

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get("DATA_DIR", "/app/data")
DB_PATH = os.path.join(DATA_DIR, "visited.db")
DIAGRAM_DIR = os.path.join(DATA_DIR, "diagrams")
LOGO_DIR = os.path.join(DATA_DIR, "logos")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(DIAGRAM_DIR, exist_ok=True)
os.makedirs(LOGO_DIR, exist_ok=True)

ALLOWED_IMAGE_EXT = {"png", "jpg", "jpeg", "webp"}

app = Flask(__name__, static_folder="static", static_url_path="")
app.config["MAX_CONTENT_LENGTH"] = 30 * 1024 * 1024  # 30MB, high-res maps can be big

LONDON_LINE_COLORS = {
    "Bakerloo": "#B36305", "Central": "#E32017", "Circle": "#FFD300", "District": "#00782A",
    "Hammersmith & City": "#F3A9BB", "Jubilee": "#A0A5A9", "Metropolitan": "#9B0056",
    "Northern": "#000000", "Piccadilly": "#003688", "Victoria": "#0098D4", "Waterloo & City": "#95CDBA",
    "DLR": "#00AFAD", "London Overground": "#FA7B05", "Elizabeth line": "#60399E",
}


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def table_exists(conn, name):
    return conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name = ?", (name,)
    ).fetchone() is not None


def column_exists(conn, table, col):
    return col in [r["name"] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]


def ensure_city(conn, city_id, name):
    exists = conn.execute("SELECT id FROM cities WHERE id = ?", (city_id,)).fetchone()
    if not exists:
        conn.execute(
            "INSERT INTO cities (id, name, created_at) VALUES (?, ?, ?)",
            (city_id, name, datetime.now(timezone.utc).isoformat()),
        )


def init_db():
    conn = get_db()

    # Pre-multi-city tables (v1.x) lack a city_id column. Rename them out of
    # the way so the names are free for the new per-city schema, then copy
    # their data across scoped to a default "london" city.
    for table in ("visited", "visited_lines", "pins"):
        if table_exists(conn, table) and not column_exists(conn, table, "city_id"):
            conn.execute(f"ALTER TABLE {table} RENAME TO {table}_legacy")
    conn.commit()

    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS cities (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            created_at TEXT NOT NULL,
            pin_size_pct REAL NOT NULL DEFAULT 0.015,
            app_title TEXT
        );
        CREATE TABLE IF NOT EXISTS lines (
            city_id TEXT NOT NULL,
            name TEXT NOT NULL,
            color TEXT NOT NULL,
            PRIMARY KEY (city_id, name)
        );
        CREATE TABLE IF NOT EXISTS stations (
            id TEXT NOT NULL,
            city_id TEXT NOT NULL,
            name TEXT NOT NULL,
            zone TEXT,
            lat REAL,
            lon REAL,
            PRIMARY KEY (id, city_id)
        );
        CREATE TABLE IF NOT EXISTS station_lines (
            station_id TEXT NOT NULL,
            city_id TEXT NOT NULL,
            line TEXT NOT NULL,
            PRIMARY KEY (station_id, city_id, line)
        );
        CREATE TABLE IF NOT EXISTS visited (
            station_id TEXT NOT NULL,
            city_id TEXT NOT NULL,
            visited_at TEXT NOT NULL,
            note TEXT,
            PRIMARY KEY (station_id, city_id)
        );
        CREATE TABLE IF NOT EXISTS visited_lines (
            station_id TEXT NOT NULL,
            city_id TEXT NOT NULL,
            line TEXT NOT NULL,
            visited_at TEXT NOT NULL,
            PRIMARY KEY (station_id, city_id, line)
        );
        """
    )
    conn.commit()

    # A station can have more than one pin on a diagram (some interchange
    # stations are drawn as separate dots per line, e.g. Ealing Broadway;
    # others as one combined dot, e.g. Blackfriars). Detect the older
    # one-pin-per-station schema (station_id+city_id as the primary key, no
    # synthetic id) and migrate it to allow multiple pins per station.
    if table_exists(conn, "pins") and not column_exists(conn, "pins", "id"):
        conn.execute("ALTER TABLE pins RENAME TO pins_single_legacy")
        conn.commit()

    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS pins (
            id TEXT PRIMARY KEY,
            station_id TEXT NOT NULL,
            city_id TEXT NOT NULL,
            x_pct REAL NOT NULL,
            y_pct REAL NOT NULL,
            lines TEXT
        );
        """
    )
    conn.commit()

    if table_exists(conn, "pins_single_legacy"):
        already = conn.execute("SELECT COUNT(*) c FROM pins").fetchone()["c"] > 0
        if not already:
            for r in conn.execute(
                "SELECT station_id, city_id, x_pct, y_pct FROM pins_single_legacy"
            ).fetchall():
                conn.execute(
                    "INSERT INTO pins (id, station_id, city_id, x_pct, y_pct, lines) VALUES (?, ?, ?, ?, ?, NULL)",
                    (str(uuid.uuid4()), r["station_id"], r["city_id"], r["x_pct"], r["y_pct"]),
                )
            conn.commit()

    if table_exists(conn, "cities") and not column_exists(conn, "cities", "pin_size_pct"):
        conn.execute("ALTER TABLE cities ADD COLUMN pin_size_pct REAL NOT NULL DEFAULT 0.015")
        conn.commit()

    if table_exists(conn, "cities") and not column_exists(conn, "cities", "app_title"):
        conn.execute("ALTER TABLE cities ADD COLUMN app_title TEXT")
        conn.commit()

    # Copy legacy (pre-multi-city) data into the new schema, once.
    if table_exists(conn, "visited_legacy"):
        already = conn.execute(
            "SELECT COUNT(*) c FROM visited WHERE city_id = 'london'"
        ).fetchone()["c"] > 0
        if not already:
            ensure_city(conn, "london", "London")
            for r in conn.execute("SELECT station_id, visited_at, note FROM visited_legacy").fetchall():
                conn.execute(
                    "INSERT OR IGNORE INTO visited (station_id, city_id, visited_at, note) VALUES (?, 'london', ?, ?)",
                    (r["station_id"], r["visited_at"], r["note"]),
                )
            if table_exists(conn, "visited_lines_legacy"):
                for r in conn.execute("SELECT station_id, line, visited_at FROM visited_lines_legacy").fetchall():
                    conn.execute(
                        "INSERT OR IGNORE INTO visited_lines (station_id, city_id, line, visited_at) VALUES (?, 'london', ?, ?)",
                        (r["station_id"], r["line"], r["visited_at"]),
                    )
            if table_exists(conn, "pins_legacy"):
                for r in conn.execute("SELECT station_id, x_pct, y_pct FROM pins_legacy").fetchall():
                    conn.execute(
                        "INSERT INTO pins (id, station_id, city_id, x_pct, y_pct, lines) VALUES (?, ?, 'london', ?, ?, NULL)",
                        (str(uuid.uuid4()), r["station_id"], r["x_pct"], r["y_pct"]),
                    )
            conn.commit()

    # Migrate the single legacy diagram file (diagrams/diagram.<ext>) to be
    # city-scoped (diagrams/london.<ext>).
    if os.path.isdir(DIAGRAM_DIR):
        for fname in os.listdir(DIAGRAM_DIR):
            if fname.startswith("diagram.") and fname.rsplit(".", 1)[-1].lower() in ALLOWED_IMAGE_EXT:
                ext = fname.rsplit(".", 1)[-1]
                os.rename(os.path.join(DIAGRAM_DIR, fname), os.path.join(DIAGRAM_DIR, f"london.{ext}"))

    # Seed London's station list from the bundled dataset if it isn't
    # populated yet (fresh install, or first boot after this migration).
    count = conn.execute("SELECT COUNT(*) c FROM stations WHERE city_id = 'london'").fetchone()["c"]
    if count == 0:
        ensure_city(conn, "london", "London")
        for line, color in LONDON_LINE_COLORS.items():
            conn.execute(
                "INSERT OR IGNORE INTO lines (city_id, name, color) VALUES ('london', ?, ?)", (line, color)
            )
        stations_path = os.path.join(APP_DIR, "static", "stations.json")
        if os.path.exists(stations_path):
            with open(stations_path) as f:
                stations = json.load(f)
            for s in stations:
                conn.execute(
                    "INSERT OR IGNORE INTO stations (id, city_id, name, zone, lat, lon) VALUES (?, 'london', ?, ?, ?, ?)",
                    (s["id"], s["name"], s.get("zone", ""), s.get("lat"), s.get("lon")),
                )
                for line in s["lines"]:
                    conn.execute(
                        "INSERT OR IGNORE INTO station_lines (station_id, city_id, line) VALUES (?, 'london', ?)",
                        (s["id"], line),
                    )
        conn.commit()

    conn.close()


init_db()


def current_diagram_path(city_id):
    return _find_city_image(DIAGRAM_DIR, city_id)


def current_logo_path(city_id):
    return _find_city_image(LOGO_DIR, city_id)


def _find_city_image(directory, city_id):
    if not os.path.isdir(directory):
        return None
    for fname in os.listdir(directory):
        base, ext = (fname.rsplit(".", 1) + [""])[:2]
        if base == city_id and ext.lower() in ALLOWED_IMAGE_EXT:
            return os.path.join(directory, fname)
    return None


def require_city(conn, city_id):
    city = conn.execute("SELECT id FROM cities WHERE id = ?", (city_id,)).fetchone()
    if not city:
        abort(404, f"unknown city '{city_id}'")


def slugify(text):
    slug = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return slug or "city"


def make_station_id(name):
    return f"{slugify(name)[:40]}-{uuid.uuid4().hex[:6]}"


@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


# --- Cities ---

@app.route("/api/cities", methods=["GET"])
def list_cities():
    conn = get_db()
    rows = conn.execute("SELECT id, name, created_at, pin_size_pct, app_title FROM cities ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/cities/<city_id>", methods=["PUT"])
def update_city(city_id):
    body = request.get_json(silent=True) or {}
    conn = get_db()
    require_city(conn, city_id)
    if "name" in body and body["name"].strip():
        conn.execute("UPDATE cities SET name = ? WHERE id = ?", (body["name"].strip(), city_id))
    if "app_title" in body:
        title = (body["app_title"] or "").strip() or None
        conn.execute("UPDATE cities SET app_title = ? WHERE id = ?", (title, city_id))
    conn.commit()
    row = conn.execute(
        "SELECT id, name, created_at, pin_size_pct, app_title FROM cities WHERE id = ?", (city_id,)
    ).fetchone()
    conn.close()
    return jsonify(dict(row))


@app.route("/api/cities/<city_id>/pin-size", methods=["PUT"])
def set_pin_size(city_id):
    body = request.get_json(silent=True) or {}
    pct = body.get("pin_size_pct")
    try:
        pct = float(pct)
    except (TypeError, ValueError):
        abort(400, "pin_size_pct must be a number")
    if not (0.001 <= pct <= 0.2):
        abort(400, "pin_size_pct must be between 0.001 and 0.2")
    conn = get_db()
    require_city(conn, city_id)
    conn.execute("UPDATE cities SET pin_size_pct = ? WHERE id = ?", (pct, city_id))
    conn.commit()
    conn.close()
    return jsonify({"id": city_id, "pin_size_pct": pct})


@app.route("/api/cities", methods=["POST"])
def create_city():
    body = request.get_json(silent=True) or {}
    name = (body.get("name") or "").strip()
    if not name:
        abort(400, "name is required")
    city_id = body.get("id") or slugify(name)
    conn = get_db()
    existing = conn.execute("SELECT id FROM cities WHERE id = ?", (city_id,)).fetchone()
    if existing:
        conn.close()
        abort(409, f"city '{city_id}' already exists")
    conn.execute(
        "INSERT INTO cities (id, name, created_at) VALUES (?, ?, ?)",
        (city_id, name, datetime.now(timezone.utc).isoformat()),
    )
    conn.commit()
    conn.close()
    return jsonify({"id": city_id, "name": name})


@app.route("/api/cities/<city_id>", methods=["DELETE"])
def delete_city(city_id):
    if city_id == "london":
        abort(400, "the default city can't be deleted")
    conn = get_db()
    require_city(conn, city_id)
    for table in ("visited_lines", "visited", "pins", "station_lines", "stations", "lines"):
        conn.execute(f"DELETE FROM {table} WHERE city_id = ?", (city_id,))
    conn.execute("DELETE FROM cities WHERE id = ?", (city_id,))
    conn.commit()
    conn.close()
    path = current_diagram_path(city_id)
    if path:
        os.remove(path)
    logo_path = current_logo_path(city_id)
    if logo_path:
        os.remove(logo_path)
    return jsonify({"deleted": True})


# --- Lines ---

@app.route("/api/cities/<city_id>/lines", methods=["GET"])
def list_lines(city_id):
    conn = get_db()
    require_city(conn, city_id)
    rows = conn.execute("SELECT name, color FROM lines WHERE city_id = ? ORDER BY name", (city_id,)).fetchall()
    conn.close()
    return jsonify({r["name"]: r["color"] for r in rows})


@app.route("/api/cities/<city_id>/lines", methods=["POST"])
def upsert_line(city_id):
    body = request.get_json(silent=True) or {}
    name = (body.get("name") or "").strip()
    color = (body.get("color") or "").strip()
    if not name or not re.match(r"^#[0-9a-fA-F]{6}$", color):
        abort(400, "name and a #RRGGBB color are required")
    conn = get_db()
    require_city(conn, city_id)
    conn.execute(
        "INSERT INTO lines (city_id, name, color) VALUES (?, ?, ?) "
        "ON CONFLICT(city_id, name) DO UPDATE SET color = excluded.color",
        (city_id, name, color),
    )
    conn.commit()
    conn.close()
    return jsonify({"name": name, "color": color})


@app.route("/api/cities/<city_id>/lines/<line>", methods=["DELETE"])
def delete_line(city_id, line):
    conn = get_db()
    require_city(conn, city_id)
    conn.execute("DELETE FROM lines WHERE city_id = ? AND name = ?", (city_id, line))
    conn.execute("DELETE FROM station_lines WHERE city_id = ? AND line = ?", (city_id, line))
    conn.commit()
    conn.close()
    return jsonify({"deleted": True})


# --- Stations ---

def _station_rows_to_json(conn, city_id, rows):
    out = []
    for r in rows:
        lines = [lr["line"] for lr in conn.execute(
            "SELECT line FROM station_lines WHERE city_id = ? AND station_id = ?", (city_id, r["id"])
        ).fetchall()]
        out.append({
            "id": r["id"], "name": r["name"], "zone": r["zone"] or "",
            "lat": r["lat"], "lon": r["lon"], "lines": lines,
        })
    return out


@app.route("/api/cities/<city_id>/stations", methods=["GET"])
def list_stations(city_id):
    conn = get_db()
    require_city(conn, city_id)
    rows = conn.execute(
        "SELECT id, name, zone, lat, lon FROM stations WHERE city_id = ? ORDER BY name", (city_id,)
    ).fetchall()
    result = _station_rows_to_json(conn, city_id, rows)
    conn.close()
    return jsonify(result)


@app.route("/api/cities/<city_id>/stations", methods=["POST"])
def create_station(city_id):
    body = request.get_json(silent=True) or {}
    name = (body.get("name") or "").strip()
    if not name:
        abort(400, "name is required")
    lines = body.get("lines") or []
    station_id = body.get("id") or make_station_id(name)
    conn = get_db()
    require_city(conn, city_id)
    existing = conn.execute(
        "SELECT id FROM stations WHERE city_id = ? AND id = ?", (city_id, station_id)
    ).fetchone()
    if existing:
        conn.close()
        abort(409, f"station id '{station_id}' already exists in this city")
    conn.execute(
        "INSERT INTO stations (id, city_id, name, zone, lat, lon) VALUES (?, ?, ?, ?, ?, ?)",
        (station_id, city_id, name, body.get("zone", ""), body.get("lat"), body.get("lon")),
    )
    for line in lines:
        conn.execute(
            "INSERT OR IGNORE INTO station_lines (station_id, city_id, line) VALUES (?, ?, ?)",
            (station_id, city_id, line),
        )
    conn.commit()
    conn.close()
    return jsonify({"id": station_id, "name": name, "zone": body.get("zone", ""),
                     "lat": body.get("lat"), "lon": body.get("lon"), "lines": lines})


@app.route("/api/cities/<city_id>/stations/<station_id>", methods=["PUT"])
def update_station(city_id, station_id):
    body = request.get_json(silent=True) or {}
    conn = get_db()
    require_city(conn, city_id)
    existing = conn.execute(
        "SELECT id FROM stations WHERE city_id = ? AND id = ?", (city_id, station_id)
    ).fetchone()
    if not existing:
        conn.close()
        abort(404, "station not found")
    conn.execute(
        "UPDATE stations SET name = ?, zone = ?, lat = ?, lon = ? WHERE city_id = ? AND id = ?",
        (body.get("name"), body.get("zone", ""), body.get("lat"), body.get("lon"), city_id, station_id),
    )
    if "lines" in body:
        conn.execute("DELETE FROM station_lines WHERE city_id = ? AND station_id = ?", (city_id, station_id))
        for line in body["lines"]:
            conn.execute(
                "INSERT OR IGNORE INTO station_lines (station_id, city_id, line) VALUES (?, ?, ?)",
                (station_id, city_id, line),
            )
    conn.commit()
    conn.close()
    return jsonify({"id": station_id, "updated": True})


@app.route("/api/cities/<city_id>/stations/<station_id>", methods=["DELETE"])
def delete_station(city_id, station_id):
    conn = get_db()
    require_city(conn, city_id)
    for table in ("visited_lines", "visited", "pins", "station_lines"):
        conn.execute(f"DELETE FROM {table} WHERE city_id = ? AND station_id = ?", (city_id, station_id))
    conn.execute("DELETE FROM stations WHERE city_id = ? AND id = ?", (city_id, station_id))
    conn.commit()
    conn.close()
    return jsonify({"deleted": True})


# --- Station bulk export / import ---

@app.route("/api/cities/<city_id>/stations/export", methods=["GET"])
def export_stations(city_id):
    fmt = request.args.get("format", "json")
    conn = get_db()
    require_city(conn, city_id)
    rows = conn.execute(
        "SELECT id, name, zone, lat, lon FROM stations WHERE city_id = ? ORDER BY name", (city_id,)
    ).fetchall()
    stations = _station_rows_to_json(conn, city_id, rows)
    conn.close()

    if fmt == "xlsx":
        if openpyxl is None:
            abort(500, "openpyxl not installed")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "stations"
        ws.append(["id", "name", "lines", "zone", "lat", "lon"])
        for s in stations:
            ws.append([s["id"], s["name"], ",".join(s["lines"]), s["zone"], s["lat"], s["lon"]])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=f"{city_id}-stations.xlsx",
        )

    buf = io.BytesIO(json.dumps(stations, indent=2).encode())
    return send_file(
        buf, mimetype="application/json", as_attachment=True, download_name=f"{city_id}-stations.json"
    )


def _parse_import_rows(file_storage):
    filename = file_storage.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "json":
        data = json.loads(file_storage.read().decode("utf-8"))
        if not isinstance(data, list):
            abort(400, "JSON import must be an array of station objects")
        rows = []
        for item in data:
            rows.append({
                "id": item.get("id"), "name": item.get("name"),
                "lines": item.get("lines") or [],
                "zone": item.get("zone", ""), "lat": item.get("lat"), "lon": item.get("lon"),
            })
        return rows

    if ext in ("xlsx", "xlsm"):
        if openpyxl is None:
            abort(500, "openpyxl not installed")
        wb = openpyxl.load_workbook(file_storage, data_only=True)
        ws = wb.active
        header = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i for i, name in enumerate(header)}
        for required in ("name",):
            if required not in idx:
                abort(400, f"Excel import is missing required column '{required}'")
        rows = []
        for row in ws.iter_rows(min_row=2):
            values = [c.value for c in row]
            if not any(values):
                continue

            def get(col):
                i = idx.get(col)
                return values[i] if i is not None and i < len(values) else None

            lines_raw = get("lines") or ""
            lines = [l.strip() for l in str(lines_raw).split(",") if l.strip()]
            rows.append({
                "id": get("id"), "name": get("name"), "lines": lines,
                "zone": get("zone") or "", "lat": get("lat"), "lon": get("lon"),
            })
        return rows

    abort(400, f"unsupported import file type: .{ext}")


@app.route("/api/cities/<city_id>/stations/import", methods=["POST"])
def import_stations(city_id):
    if "file" not in request.files:
        abort(400, "no file uploaded under field 'file'")
    rows = _parse_import_rows(request.files["file"])

    conn = get_db()
    require_city(conn, city_id)
    imported, updated, skipped, errors = 0, 0, 0, []

    for i, row in enumerate(rows):
        name = (row.get("name") or "").strip()
        if not name:
            skipped += 1
            errors.append(f"row {i + 2}: missing name")
            continue
        station_id = row.get("id")
        if not station_id:
            existing_by_name = conn.execute(
                "SELECT id FROM stations WHERE city_id = ? AND lower(name) = lower(?)", (city_id, name)
            ).fetchone()
            station_id = existing_by_name["id"] if existing_by_name else make_station_id(name)
        existing = conn.execute(
            "SELECT id FROM stations WHERE city_id = ? AND id = ?", (city_id, station_id)
        ).fetchone()
        conn.execute(
            "INSERT INTO stations (id, city_id, name, zone, lat, lon) VALUES (?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(id, city_id) DO UPDATE SET name=excluded.name, zone=excluded.zone, "
            "lat=excluded.lat, lon=excluded.lon",
            (station_id, city_id, name, row.get("zone", ""), row.get("lat"), row.get("lon")),
        )
        conn.execute("DELETE FROM station_lines WHERE city_id = ? AND station_id = ?", (city_id, station_id))
        for line in row.get("lines") or []:
            conn.execute(
                "INSERT OR IGNORE INTO station_lines (station_id, city_id, line) VALUES (?, ?, ?)",
                (station_id, city_id, line),
            )
            conn.execute(
                "INSERT OR IGNORE INTO lines (city_id, name, color) VALUES (?, ?, ?)",
                (city_id, line, "#888888"),
            )
        if existing:
            updated += 1
        else:
            imported += 1

    conn.commit()
    conn.close()
    return jsonify({"imported": imported, "updated": updated, "skipped": skipped, "errors": errors})


# --- Visited (per city) ---

@app.route("/api/cities/<city_id>/visited", methods=["GET"])
def list_visited(city_id):
    conn = get_db()
    require_city(conn, city_id)
    rows = conn.execute(
        "SELECT station_id, visited_at, note FROM visited WHERE city_id = ?", (city_id,)
    ).fetchall()
    result = {}
    for r in rows:
        lines = [lr["line"] for lr in conn.execute(
            "SELECT line FROM visited_lines WHERE city_id = ? AND station_id = ?", (city_id, r["station_id"])
        ).fetchall()]
        result[r["station_id"]] = {"visited_at": r["visited_at"], "note": r["note"], "lines": lines}
    conn.close()
    return jsonify(result)


@app.route("/api/cities/<city_id>/visited/<station_id>", methods=["POST"])
def toggle_visited(city_id, station_id):
    conn = get_db()
    require_city(conn, city_id)
    existing = conn.execute(
        "SELECT station_id FROM visited WHERE city_id = ? AND station_id = ?", (city_id, station_id)
    ).fetchone()

    if existing:
        conn.execute("DELETE FROM visited WHERE city_id = ? AND station_id = ?", (city_id, station_id))
        conn.execute("DELETE FROM visited_lines WHERE city_id = ? AND station_id = ?", (city_id, station_id))
        conn.commit()
        conn.close()
        return jsonify({"station_id": station_id, "visited": False})

    body = request.get_json(silent=True) or {}
    visited_at = datetime.now(timezone.utc).isoformat()
    conn.execute(
        "INSERT INTO visited (station_id, city_id, visited_at, note) VALUES (?, ?, ?, ?)",
        (station_id, city_id, visited_at, body.get("note")),
    )
    conn.commit()
    conn.close()
    return jsonify({"station_id": station_id, "visited": True, "visited_at": visited_at, "lines": []})


@app.route("/api/cities/<city_id>/visited/<station_id>/note", methods=["PUT"])
def set_note(city_id, station_id):
    body = request.get_json(silent=True) or {}
    conn = get_db()
    require_city(conn, city_id)
    conn.execute(
        "UPDATE visited SET note = ? WHERE city_id = ? AND station_id = ?",
        (body.get("note", ""), city_id, station_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"station_id": station_id, "note": body.get("note", "")})


def ensure_visited(conn, city_id, station_id):
    exists = conn.execute(
        "SELECT station_id FROM visited WHERE city_id = ? AND station_id = ?", (city_id, station_id)
    ).fetchone()
    if not exists:
        conn.execute(
            "INSERT INTO visited (station_id, city_id, visited_at, note) VALUES (?, ?, ?, NULL)",
            (station_id, city_id, datetime.now(timezone.utc).isoformat()),
        )


@app.route("/api/cities/<city_id>/visited/<station_id>/lines/<line>", methods=["POST"])
def add_visited_line(city_id, station_id, line):
    conn = get_db()
    require_city(conn, city_id)
    ensure_visited(conn, city_id, station_id)
    conn.execute(
        "INSERT OR IGNORE INTO visited_lines (station_id, city_id, line, visited_at) VALUES (?, ?, ?, ?)",
        (station_id, city_id, line, datetime.now(timezone.utc).isoformat()),
    )
    conn.commit()
    lines = [r["line"] for r in conn.execute(
        "SELECT line FROM visited_lines WHERE city_id = ? AND station_id = ?", (city_id, station_id)
    ).fetchall()]
    conn.close()
    return jsonify({"station_id": station_id, "lines": lines, "visited": True})


@app.route("/api/cities/<city_id>/visited/<station_id>/lines/<line>", methods=["DELETE"])
def remove_visited_line(city_id, station_id, line):
    conn = get_db()
    require_city(conn, city_id)
    conn.execute(
        "DELETE FROM visited_lines WHERE city_id = ? AND station_id = ? AND line = ?",
        (city_id, station_id, line),
    )
    conn.commit()
    lines = [r["line"] for r in conn.execute(
        "SELECT line FROM visited_lines WHERE city_id = ? AND station_id = ?", (city_id, station_id)
    ).fetchall()]
    conn.close()
    return jsonify({"station_id": station_id, "lines": lines})


# --- Diagram pins (per city). A station can have more than one pin —
# some interchange stations are drawn as separate dots per line on the
# official diagram, others as one combined dot. `lines` on a pin scopes
# it to a subset of the station's lines; NULL/empty means "the whole
# station" (the common case).

def _pin_row_to_json(r):
    lines = [l for l in (r["lines"] or "").split(",") if l]
    return {"id": r["id"], "station_id": r["station_id"], "x_pct": r["x_pct"], "y_pct": r["y_pct"], "lines": lines}


@app.route("/api/cities/<city_id>/pins", methods=["GET"])
def list_pins(city_id):
    conn = get_db()
    require_city(conn, city_id)
    rows = conn.execute(
        "SELECT id, station_id, x_pct, y_pct, lines FROM pins WHERE city_id = ?", (city_id,)
    ).fetchall()
    conn.close()
    return jsonify([_pin_row_to_json(r) for r in rows])


@app.route("/api/cities/<city_id>/pins", methods=["POST"])
def create_pin(city_id):
    body = request.get_json(silent=True) or {}
    station_id = body.get("station_id")
    x_pct, y_pct = body.get("x_pct"), body.get("y_pct")
    if not station_id or x_pct is None or y_pct is None:
        abort(400, "station_id, x_pct and y_pct are required")
    lines = body.get("lines") or []
    lines_str = ",".join(lines) if lines else None

    conn = get_db()
    require_city(conn, city_id)
    pin_id = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO pins (id, station_id, city_id, x_pct, y_pct, lines) VALUES (?, ?, ?, ?, ?, ?)",
        (pin_id, station_id, city_id, x_pct, y_pct, lines_str),
    )
    conn.commit()
    conn.close()
    return jsonify({"id": pin_id, "station_id": station_id, "x_pct": x_pct, "y_pct": y_pct, "lines": lines})


@app.route("/api/cities/<city_id>/pins/<pin_id>", methods=["PUT"])
def update_pin(city_id, pin_id):
    body = request.get_json(silent=True) or {}
    conn = get_db()
    require_city(conn, city_id)
    existing = conn.execute(
        "SELECT id FROM pins WHERE city_id = ? AND id = ?", (city_id, pin_id)
    ).fetchone()
    if not existing:
        conn.close()
        abort(404, "pin not found")

    if "x_pct" in body and "y_pct" in body:
        conn.execute(
            "UPDATE pins SET x_pct = ?, y_pct = ? WHERE city_id = ? AND id = ?",
            (body["x_pct"], body["y_pct"], city_id, pin_id),
        )
    if "lines" in body:
        lines_str = ",".join(body["lines"]) if body["lines"] else None
        conn.execute("UPDATE pins SET lines = ? WHERE city_id = ? AND id = ?", (lines_str, city_id, pin_id))
    conn.commit()
    row = conn.execute(
        "SELECT id, station_id, x_pct, y_pct, lines FROM pins WHERE city_id = ? AND id = ?", (city_id, pin_id)
    ).fetchone()
    conn.close()
    return jsonify(_pin_row_to_json(row))


@app.route("/api/cities/<city_id>/pins/<pin_id>", methods=["DELETE"])
def delete_pin(city_id, pin_id):
    conn = get_db()
    require_city(conn, city_id)
    conn.execute("DELETE FROM pins WHERE city_id = ? AND id = ?", (city_id, pin_id))
    conn.commit()
    conn.close()
    return jsonify({"id": pin_id, "deleted": True})


# --- Diagram image (per city) ---

@app.route("/api/cities/<city_id>/diagram", methods=["GET"])
def diagram_status(city_id):
    conn = get_db()
    require_city(conn, city_id)
    conn.close()
    return jsonify({"exists": current_diagram_path(city_id) is not None})


@app.route("/api/cities/<city_id>/diagram", methods=["POST"])
def upload_diagram(city_id):
    conn = get_db()
    require_city(conn, city_id)
    conn.close()
    if "image" not in request.files:
        abort(400, "no file uploaded under field 'image'")
    file = request.files["image"]
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED_IMAGE_EXT:
        abort(400, f"unsupported file type .{ext}")
    old = current_diagram_path(city_id)
    if old:
        os.remove(old)
    safe_name = secure_filename(f"{city_id}.{ext}")
    file.save(os.path.join(DIAGRAM_DIR, safe_name))
    return jsonify({"ok": True})


@app.route("/api/cities/<city_id>/diagram", methods=["DELETE"])
def delete_diagram(city_id):
    path = current_diagram_path(city_id)
    if path:
        os.remove(path)
    return jsonify({"deleted": True})


@app.route("/api/cities/<city_id>/diagram/image", methods=["GET"])
def diagram_image(city_id):
    path = current_diagram_path(city_id)
    if not path:
        abort(404)
    mime = mimetypes.guess_type(path)[0] or "application/octet-stream"
    return send_file(path, mimetype=mime)


# --- Per-city logo (also used as the browser tab favicon) ---

@app.route("/api/cities/<city_id>/logo", methods=["GET"])
def logo_status(city_id):
    conn = get_db()
    require_city(conn, city_id)
    conn.close()
    return jsonify({"exists": current_logo_path(city_id) is not None})


@app.route("/api/cities/<city_id>/logo", methods=["POST"])
def upload_logo(city_id):
    conn = get_db()
    require_city(conn, city_id)
    conn.close()
    if "image" not in request.files:
        abort(400, "no file uploaded under field 'image'")
    file = request.files["image"]
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED_IMAGE_EXT:
        abort(400, f"unsupported file type .{ext}")
    old = current_logo_path(city_id)
    if old:
        os.remove(old)
    safe_name = secure_filename(f"{city_id}.{ext}")
    file.save(os.path.join(LOGO_DIR, safe_name))
    return jsonify({"ok": True})


@app.route("/api/cities/<city_id>/logo", methods=["DELETE"])
def delete_logo(city_id):
    path = current_logo_path(city_id)
    if path:
        os.remove(path)
    return jsonify({"deleted": True})


@app.route("/api/cities/<city_id>/logo/image", methods=["GET"])
def logo_image(city_id):
    path = current_logo_path(city_id)
    if not path:
        abort(404)
    mime = mimetypes.guess_type(path)[0] or "application/octet-stream"
    return send_file(path, mimetype=mime)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000)
